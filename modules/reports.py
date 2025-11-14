from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
                             QTableWidget, QTableWidgetItem, QComboBox, QDateEdit, QGroupBox,
                             QHeaderView, QFrame, QMessageBox)
from PyQt5.QtCore import QDate, Qt
from utils.formatters import format_currency
from modules.report_components import StatsCard, DataHelper, TableManager

class ReportsModule:
    def __init__(self, db):
        self.db = db
        self.widget = QWidget()
        self.init_ui()
        
    def get_widget(self):
        return self.widget
        
    def on_enter(self):
        self.load_reports_data()
        
    def on_leave(self):
        pass
        
    def init_ui(self):
        layout = QVBoxLayout(self.widget)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(20)
        
        self.create_header(layout)
        self.create_date_filters(layout)
        self.create_main_stats(layout)
        self.create_detailed_reports(layout)
        
    def create_header(self, layout):
        header_layout = QHBoxLayout()
        title = QLabel("📊 Reportes del Kiosco")
        title.setStyleSheet("font-size: 28px; font-weight: bold; color: #1e293b;")
        
        export_btn = QPushButton("📤 Exportar Reporte")
        export_btn.setProperty("class", "success")
        export_btn.clicked.connect(self.export_report)
        
        header_layout.addWidget(title)
        header_layout.addStretch()
        header_layout.addWidget(export_btn)
        layout.addLayout(header_layout)
        
    def create_date_filters(self, layout):
        filter_frame = QFrame()
        filter_frame.setStyleSheet("QFrame { background-color: white; border-radius: 12px; padding: 20px; }")
        filter_layout = QHBoxLayout(filter_frame)
    
        filter_layout.addWidget(QLabel("Período:"))
        self.period_combo = QComboBox()
        self.period_combo.addItems(["Hoy", "Ayer", "Últimos 7 días", "Este mes", "Mes anterior", "Personalizado"])
        self.period_combo.currentTextChanged.connect(self.on_period_change)
    
        self.start_date = QDateEdit()
        self.start_date.setDate(QDate.currentDate().addDays(-7))
        self.start_date.setCalendarPopup(True)
        self.start_date.setDisplayFormat("dd/MM/yyyy")
    
        self.end_date = QDateEdit()
        self.end_date.setDate(QDate.currentDate())
        self.end_date.setCalendarPopup(True)
        self.end_date.setDisplayFormat("dd/MM/yyyy")
    
        # Agregar opción de vista
        self.view_combo = QComboBox()
        self.view_combo.addItems(["📊 Resumen", "📋 Detallado"])
        self.view_combo.setToolTip("Cambiar entre vista resumen o detallada")

        update_btn = QPushButton("🔄 Actualizar Reporte")
        update_btn.setStyleSheet("background-color: #3b82f6; color: white; font-weight: bold; padding: 10px 20px;")
        update_btn.clicked.connect(self.load_reports_data)

        filter_layout.addWidget(self.period_combo)
        filter_layout.addWidget(QLabel("Vista:"))
        filter_layout.addWidget(self.view_combo)
        filter_layout.addWidget(QLabel("Desde:"))
        filter_layout.addWidget(self.start_date)
        filter_layout.addWidget(QLabel("Hasta:"))
        filter_layout.addWidget(self.end_date)
        filter_layout.addWidget(update_btn)
        filter_layout.addStretch()
    
        layout.addWidget(filter_frame)
        
    def create_main_stats(self, layout):
        self.stats_layout = QHBoxLayout()
        self.stats_layout.setSpacing(16)
        
        self.stats_frames = []
        stats_templates = [
            ("Ventas Totales", "$ 0", "+0%", "#3b82f6", "💰", "Cargando..."),
            ("Productos Vendidos", "0", "+0%", "#10b981", "📦", "unidades"),
            ("Ticket Promedio", "$ 0", "+0%", "#f59e0b", "🎫", "por venta"),
            ("Clientes Atendidos", "0", "+0%", "#ef4444", "👥", "personas"),
        ]
        
        for text, value, trend, color, icon, subtitle in stats_templates:
            stat_frame = StatsCard.create(text, value, trend, color, icon, subtitle)
            self.stats_layout.addWidget(stat_frame)
            self.stats_frames.append(stat_frame)
        
        layout.addLayout(self.stats_layout)
        
    def create_detailed_reports(self, layout):
        columns_layout = QHBoxLayout()
        columns_layout.setSpacing(20)

        # Columna izquierda
        left_column = QVBoxLayout()
        left_column.setSpacing(16)

        self.hourly_group = QGroupBox("📈 Ventas por Horario")
        self.hourly_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        self.hourly_layout = QVBoxLayout(self.hourly_group)
        left_column.addWidget(self.hourly_group)

        self.payment_group = QGroupBox("💳 Distribución de Pagos")
        self.payment_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        self.payment_layout = QVBoxLayout(self.payment_group)
        left_column.addWidget(self.payment_group)

        # Columna derecha
        right_column = QVBoxLayout()
        right_column.setSpacing(16)

        # Top productos
        products_group = QGroupBox("🏆 Productos Más Vendidos")
        products_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        products_layout = QVBoxLayout(products_group)
        self.top_products_table = QTableWidget()
        TableManager.setup_products_table(self.top_products_table)
        products_layout.addWidget(self.top_products_table)
        right_column.addWidget(products_group)

        # Ventas recientes
        recent_group = QGroupBox("🔄 Ventas Recientes")
        recent_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        recent_layout = QVBoxLayout(recent_group)
        self.recent_sales_table = QTableWidget()
        TableManager.setup_sales_table(self.recent_sales_table)
        recent_layout.addWidget(self.recent_sales_table)
        right_column.addWidget(recent_group)

        columns_layout.addLayout(left_column, 1)
        columns_layout.addLayout(right_column, 2)
        layout.addLayout(columns_layout)

    def on_period_change(self, period):
        today = QDate.currentDate()
        period_config = {
            "Hoy": (today, today),
            "Ayer": (today.addDays(-1), today.addDays(-1)),
            "Últimos 7 días": (today.addDays(-7), today),
            "Este mes": (QDate(today.year(), today.month(), 1), today),
            "Mes anterior": (
                QDate(today.year(), today.month()-1, 1),
                QDate(today.year(), today.month(), 1).addDays(-1)
            )
        }
        
        if period in period_config:
            start, end = period_config[period]
            self.start_date.setDate(start)
            self.end_date.setDate(end)
            
    def load_reports_data(self):
        """Cargar datos de reportes desde la base de datos"""
        try:
            start_date = self.start_date.date().toString("yyyy-MM-dd")
            end_date = self.end_date.date().toString("yyyy-MM-dd")
            view_mode = self.view_combo.currentText()

            # Determinar si mostrar vista detallada o resumen
            is_detailed = "📋 Detallado" in view_mode

            # Cargar datos reales
            current_sales_summary = self.db.get_sales_summary(start_date, end_date)
            previous_sales_summary = self.db.get_previous_period_sales(start_date, end_date)
            top_products = self.db.get_top_products(start_date, end_date, 10)
            recent_sales = self.db.get_sales_report(start_date, end_date)
            current_products_sold = self.db.get_total_products_sold(start_date, end_date)

            # Calcular período anterior
            prev_start, prev_end = self.get_previous_period_dates(start_date, end_date)
            previous_products_sold = self.db.get_total_products_sold(prev_start, prev_end)

            # Actualizar componentes
            self.update_main_stats(current_sales_summary, previous_sales_summary,
                                 current_products_sold, previous_products_sold)
            self.update_top_products(top_products)
            self.update_recent_sales(recent_sales)
            self.update_hourly_sales(start_date, end_date)
            self.update_payment_methods(start_date, end_date)

            # Mostrar/ocultar secciones según la vista seleccionada
            self.toggle_view_mode(is_detailed)

            # Show success message only if data was loaded successfully
            if current_sales_summary or top_products or recent_sales:
                QMessageBox.information(self.widget, "✅ Actualizado", f"Reportes actualizados correctamente ({view_mode})")
            else:
                QMessageBox.information(self.widget, "ℹ️ Sin Datos", "No hay datos para el período seleccionado.")

        except Exception as e:
            QMessageBox.critical(self.widget, "❌ Error", f"Error al cargar reportes:\n{str(e)}")
            
    def update_main_stats(self, current_sales_summary, previous_sales_summary, 
                        current_products_sold, previous_products_sold):
        """Actualizar estadísticas principales"""
        try:
            # Obtener datos actuales y anteriores
            current_data = current_sales_summary or (0, 0, 0, 0)
            previous_data = previous_sales_summary or (0, 0, 0, 0)
            
            current_sales, current_amount, current_avg_ticket, current_customers = current_data
            previous_sales, previous_amount, previous_avg_ticket, previous_customers = previous_data

            # Calcular métricas
            stats_data = [
                (format_currency(current_amount), 
                 DataHelper.calculate_growth(current_amount, previous_amount), 
                 f"{current_sales} ventas"),
                (str(current_products_sold), 
                 DataHelper.calculate_growth(current_products_sold, previous_products_sold), 
                 "unidades vendidas"),
                (format_currency(current_avg_ticket), 
                 DataHelper.calculate_growth(current_avg_ticket, previous_avg_ticket), 
                 "ticket promedio"),
                (str(current_customers), 
                 DataHelper.calculate_growth(current_customers, previous_customers), 
                 "clientes únicos")
            ]

            # Actualizar UI
            for i, (value, growth, subtitle) in enumerate(stats_data):
                if i < len(self.stats_frames):
                    frame = self.stats_frames[i]
                    color = DataHelper.get_growth_color(growth)
                    
                    # Buscar componentes
                    value_label = frame.findChild(QLabel, "value")
                    trend_label = frame.findChild(QLabel, "trend")
                    subtitle_label = frame.findChild(QLabel, "subtitle")
                    
                    if value_label: value_label.setText(str(value))
                    if trend_label: 
                        trend_label.setText(growth)
                        trend_label.setStyleSheet(f"color: {color}; font-size: 12px; font-weight: 600;")
                    if subtitle_label: subtitle_label.setText(subtitle)
                    
        except Exception as e:
            print(f"Error en update_main_stats: {e}")
                    
    def update_top_products(self, top_products):
        """Actualizar tabla de productos más vendidos"""
        if not top_products:
            self.top_products_table.setRowCount(0)
            return
            
        self.top_products_table.setRowCount(len(top_products))
        
        for row, (name, quantity, amount) in enumerate(top_products):
            name = name or "Producto sin nombre"
            quantity = quantity or 0
            amount = amount or 0.0
            
            self.top_products_table.setItem(row, 0, QTableWidgetItem(str(row + 1)))
            self.top_products_table.setItem(row, 1, QTableWidgetItem(str(name)))
            self.top_products_table.setItem(row, 2, QTableWidgetItem(str(int(quantity))))
            self.top_products_table.setItem(row, 3, QTableWidgetItem(format_currency(amount)))
            
    def update_recent_sales(self, recent_sales):
        """Actualizar tabla de ventas recientes"""
        try:
            if not recent_sales or not isinstance(recent_sales, (list, tuple)):
                self.recent_sales_table.setRowCount(0)
                return
                
            self.recent_sales_table.setRowCount(len(recent_sales))
            
            for row, sale in enumerate(recent_sales):
                if not isinstance(sale, (tuple, list)) or len(sale) < 5:
                    continue
                    
                # Extraer datos
                sale_id, total, payment_method, customer_type, created_at = sale[:5]
                items_count = sale[5] if len(sale) > 5 else 0
                
                # Formatear datos
                sale_time = DataHelper.adjust_timezone(str(created_at))
                items_description = f"{items_count} productos" if items_count else "Sin detalles"
                
                # Llenar tabla
                self.recent_sales_table.setItem(row, 0, QTableWidgetItem(sale_time))
                self.recent_sales_table.setItem(row, 1, QTableWidgetItem(items_description))
                self.recent_sales_table.setItem(row, 2, QTableWidgetItem(format_currency(total)))
                self.recent_sales_table.setItem(row, 3, QTableWidgetItem(str(payment_method)))
                self.recent_sales_table.setItem(row, 4, QTableWidgetItem(str(customer_type)))
                
        except Exception as e:
            print(f"Error en update_recent_sales: {e}")
            self.recent_sales_table.setRowCount(0)

    def update_hourly_sales(self, start_date, end_date):
        """Actualizar ventas por horario mostrando las ventas individuales"""
        from PyQt5.QtWidgets import QWidget, QHBoxLayout, QLabel

        # Limpiar layout
        while self.hourly_layout.count():
            child = self.hourly_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()

        try:
            # Obtener ventas detalladas del período
            detailed_sales = self.db.get_detailed_sales_report(start_date, end_date)

            if not detailed_sales:
                # Si no hay datos, mostrar mensaje
                no_data_widget = QWidget()
                no_data_layout = QHBoxLayout(no_data_widget)
                no_data_label = QLabel("No hay ventas en el período seleccionado")
                no_data_label.setStyleSheet("color: #6b7280; font-style: italic;")
                no_data_layout.addWidget(no_data_label)
                self.hourly_layout.addWidget(no_data_widget)
                return

            # Mostrar cada venta individual
            for sale in detailed_sales:
                sale_id, total, payment_method, customer_type, created_at, payment_status, items_detail, items_count = sale

                # Formatear hora
                sale_time = DataHelper.adjust_timezone(str(created_at))
                time_only = sale_time.split(' ')[1] if ' ' in sale_time else sale_time

                # Formatear detalles
                amount = format_currency(total)
                customer = customer_type if customer_type != 'Consumidor Final' else 'Final'
                details = f"{items_count} productos - {customer}"

                # Mostrar la venta
                self.add_hourly_item(time_only, amount, details, payment_method)

        except Exception as e:
            print(f"Error en update_hourly_sales: {e}")
            # Mostrar mensaje de error
            error_widget = QWidget()
            error_layout = QHBoxLayout(error_widget)
            error_label = QLabel("Error al cargar datos de ventas por horario")
            error_label.setStyleSheet("color: #ef4444; font-style: italic;")
            error_layout.addWidget(error_label)
            self.hourly_layout.addWidget(error_widget)

    def add_hourly_item(self, time, amount, details, percentage):
        time_widget = QWidget()
        time_layout = QHBoxLayout(time_widget)
        time_layout.setContentsMargins(8, 6, 8, 6)
        
        time_label = QLabel(time)
        time_label.setStyleSheet("font-weight: 600; color: #374151;")
        
        amount_label = QLabel(amount)
        amount_label.setStyleSheet("font-weight: bold; color: #059669;")
        
        details_label = QLabel(details)
        details_label.setStyleSheet("color: #6b7280; font-size: 12px;")
        
        percentage_label = QLabel(percentage)
        percentage_label.setStyleSheet("color: #10b981; font-weight: 600;" if percentage.startswith('+') else "color: #ef4444; font-weight: 600;")
        
        time_layout.addWidget(time_label)
        time_layout.addStretch()
        time_layout.addWidget(details_label)
        time_layout.addWidget(amount_label)
        time_layout.addWidget(percentage_label)
        
        self.hourly_layout.addWidget(time_widget)

    def update_payment_methods(self, start_date, end_date):
        """Actualizar distribución de métodos de pago con datos reales"""
        from PyQt5.QtWidgets import QWidget, QHBoxLayout, QLabel

        # Limpiar layout
        while self.payment_layout.count():
            child = self.payment_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()

        try:
            # Obtener datos reales de la base de datos
            payment_data = self.db.get_payment_methods_distribution(start_date, end_date)

            if not payment_data:
                # Si no hay datos, mostrar mensaje
                no_data_widget = QWidget()
                no_data_layout = QHBoxLayout(no_data_widget)
                no_data_label = QLabel("No hay datos de pagos en el período seleccionado")
                no_data_label.setStyleSheet("color: #6b7280; font-style: italic;")
                no_data_layout.addWidget(no_data_label)
                self.payment_layout.addWidget(no_data_widget)
                return

            # Calcular total para porcentajes
            total_amount = sum(amount for _, _, amount in payment_data)

            # Mostrar métodos de pago
            for method, count, amount in payment_data:
                if amount > 0:  # Solo mostrar métodos con transacciones
                    percentage = f"{(amount / total_amount * 100):.1f}%" if total_amount > 0 else "0%"
                    formatted_amount = format_currency(amount)
                    color = "#10b981"  # Verde por defecto
                    self.add_payment_item(method, formatted_amount, percentage, color)

        except Exception as e:
            print(f"Error en update_payment_methods: {e}")
            # Mostrar mensaje de error
            error_widget = QWidget()
            error_layout = QHBoxLayout(error_widget)
            error_label = QLabel("Error al cargar distribución de pagos")
            error_label.setStyleSheet("color: #ef4444; font-style: italic;")
            error_layout.addWidget(error_label)
            self.payment_layout.addWidget(error_widget)

    def add_payment_item(self, method, amount, percentage, color):
        method_widget = QWidget()
        method_layout = QHBoxLayout(method_widget)
        method_layout.setContentsMargins(8, 8, 8, 8)
        
        method_label = QLabel(method)
        method_label.setStyleSheet("font-weight: 600; min-width: 100px;")
        
        amount_label = QLabel(amount)
        amount_label.setStyleSheet("font-weight: bold; color: #374151;")
        
        percentage_label = QLabel(percentage)
        percentage_label.setStyleSheet("font-weight: 600; color: #6b7280; min-width: 40px;")
        
        method_layout.addWidget(method_label)
        method_layout.addStretch()
        method_layout.addWidget(amount_label)
        method_layout.addWidget(percentage_label)
        
        self.payment_layout.addWidget(method_widget)

    def update_debug_info(self):
        """Actualizar información de debug de la base de datos"""
        try:
            import sqlite3
            conn = sqlite3.connect('kiosco_pos.db')
            cursor = conn.cursor()

            # Ver tablas
            cursor.execute('SELECT name FROM sqlite_master WHERE type="table"')
            tables = cursor.fetchall()
            debug_info = "Tablas en la base de datos:\n"
            for table in tables:
                debug_info += f"- {table[0]}\n"

            # Ver estructura de sales
            debug_info += "\nEstructura de tabla sales:\n"
            cursor.execute('PRAGMA table_info(sales)')
            columns = cursor.fetchall()
            for col in columns:
                debug_info += f"  {col[1]}: {col[2]}\n"

            # Ver estructura de sale_items
            debug_info += "\nEstructura de tabla sale_items:\n"
            cursor.execute('PRAGMA table_info(sale_items)')
            columns = cursor.fetchall()
            for col in columns:
                debug_info += f"  {col[1]}: {col[2]}\n"

            # Ver últimas ventas
            debug_info += "\nÚltimas 5 ventas:\n"
            cursor.execute('SELECT id, total, payment_method, payment_status, created_at FROM sales ORDER BY created_at DESC LIMIT 5')
            sales = cursor.fetchall()
            for sale in sales:
                debug_info += f"  ID: {sale[0]}, Total: {sale[1]}, Método: {sale[2]}, Estado: {sale[3]}, Fecha: {sale[4]}\n"

            # Ver items de las últimas ventas
            debug_info += "\nItems de las últimas ventas:\n"
            cursor.execute('''
                SELECT si.sale_id, si.product_name, si.quantity, si.price, si.subtotal
                FROM sale_items si
                JOIN sales s ON si.sale_id = s.id
                ORDER BY s.created_at DESC LIMIT 10
            ''')
            items = cursor.fetchall()
            for item in items:
                debug_info += f"  Venta {item[0]}: {item[1]} x{item[2]} = {item[4]}\n"

            conn.close()
            self.debug_text.setText(debug_info)

        except Exception as e:
            self.debug_text.setText(f"Error al cargar información de debug:\n{str(e)}")

    def export_report(self):
        """Exportar reporte a Excel"""
        try:
            from PyQt5.QtWidgets import QFileDialog, QMessageBox
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from datetime import datetime

            # Obtener fechas del período actual
            start_date = self.start_date.date().toString("yyyy-MM-dd")
            end_date = self.end_date.date().toString("yyyy-MM-dd")

            # Abrir diálogo para guardar archivo
            filename, _ = QFileDialog.getSaveFileName(
                self.widget,
                "Exportar Reporte",
                f"reporte_{start_date}_a_{end_date}.xlsx",
                "Archivos Excel (*.xlsx)"
            )

            if not filename:
                return

            # Crear workbook de Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte de Ventas"

            # Estilos
            header_font = Font(bold=True, size=14, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            subheader_font = Font(bold=True, size=12, color="000000")
            subheader_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
            data_font = Font(size=10)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            # Fila actual para escribir
            current_row = 1

            # Título principal
            ws.merge_cells(f'A{current_row}:F{current_row}')
            title_cell = ws[f'A{current_row}']
            title_cell.value = "📊 REPORTE DE VENTAS - KIOSCO POS"
            title_cell.font = Font(bold=True, size=16, color="1e293b")
            title_cell.alignment = Alignment(horizontal="center")
            current_row += 1

            # Información del período
            ws.merge_cells(f'A{current_row}:F{current_row}')
            period_cell = ws[f'A{current_row}']
            period_cell.value = f"Período: {start_date} a {end_date}"
            period_cell.font = Font(bold=True, size=12)
            period_cell.alignment = Alignment(horizontal="center")
            current_row += 1

            # Fecha de generación
            ws.merge_cells(f'A{current_row}:F{current_row}')
            date_cell = ws[f'A{current_row}']
            date_cell.value = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            date_cell.font = Font(italic=True, size=10)
            date_cell.alignment = Alignment(horizontal="center")
            current_row += 2

            # Obtener datos para exportar
            sales_data = self.db.get_sales_report(start_date, end_date)
            top_products = self.db.get_top_products(start_date, end_date, 50)
            sales_summary = self.db.get_sales_summary(start_date, end_date)
            payment_data = self.db.get_payment_methods_distribution(start_date, end_date)
            detailed_sales = self.db.get_detailed_sales_report(start_date, end_date)

            # Resumen general
            if sales_summary:
                # Título de sección
                ws.merge_cells(f'A{current_row}:F{current_row}')
                summary_title = ws[f'A{current_row}']
                summary_title.value = "RESUMEN GENERAL"
                summary_title.font = subheader_font
                summary_title.fill = subheader_fill
                summary_title.alignment = Alignment(horizontal="center")
                current_row += 1

                sales_count, total_amount, avg_ticket, customers = sales_summary

                # Encabezados
                headers = ['Métrica', 'Valor']
                for col, header in enumerate(headers, 1):
                    cell = ws[f'{get_column_letter(col)}{current_row}']
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = border
                current_row += 1

                # Datos del resumen
                summary_data = [
                    ['Ventas Totales', sales_count],
                    ['Monto Total', format_currency(total_amount)],
                    ['Ticket Promedio', format_currency(avg_ticket)],
                    ['Clientes Atendidos', customers]
                ]

                for row_data in summary_data:
                    for col, value in enumerate(row_data, 1):
                        cell = ws[f'{get_column_letter(col)}{current_row}']
                        cell.value = value
                        cell.font = data_font
                        cell.border = border
                        if col == 1:
                            cell.alignment = Alignment(horizontal="left")
                        else:
                            cell.alignment = Alignment(horizontal="right")
                    current_row += 1

                current_row += 1

            # Ventas detalladas
            if sales_data:
                # Título de sección
                ws.merge_cells(f'A{current_row}:F{current_row}')
                sales_title = ws[f'A{current_row}']
                sales_title.value = "VENTAS DETALLADAS"
                sales_title.font = subheader_font
                sales_title.fill = subheader_fill
                sales_title.alignment = Alignment(horizontal="center")
                current_row += 1

                # Encabezados
                headers = ['ID Venta', 'Fecha/Hora', 'Total', 'Método de Pago', 'Tipo Cliente', 'Items']
                for col, header in enumerate(headers, 1):
                    cell = ws[f'{get_column_letter(col)}{current_row}']
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = border
                current_row += 1

                # Datos de ventas
                for sale in sales_data:
                    if len(sale) >= 5:
                        sale_id, total, payment_method, customer_type, created_at = sale[:5]
                        items_count = sale[5] if len(sale) > 5 else 0

                        sale_time = DataHelper.adjust_timezone(str(created_at))

                        row_data = [
                            sale_id,
                            sale_time,
                            format_currency(total),
                            payment_method,
                            customer_type,
                            items_count
                        ]

                        for col, value in enumerate(row_data, 1):
                            cell = ws[f'{get_column_letter(col)}{current_row}']
                            cell.value = value
                            cell.font = data_font
                            cell.border = border
                            if col in [3]:  # Columnas numéricas
                                cell.alignment = Alignment(horizontal="right")
                            elif col in [1, 6]:  # ID e Items
                                cell.alignment = Alignment(horizontal="center")
                            else:
                                cell.alignment = Alignment(horizontal="left")
                        current_row += 1

                current_row += 1

            # Productos más vendidos
            if top_products:
                # Título de sección
                ws.merge_cells(f'A{current_row}:C{current_row}')
                products_title = ws[f'A{current_row}']
                products_title.value = "PRODUCTOS MÁS VENDIDOS"
                products_title.font = subheader_font
                products_title.fill = subheader_fill
                products_title.alignment = Alignment(horizontal="center")
                current_row += 1

                # Encabezados
                headers = ['Producto', 'Cantidad', 'Monto Total']
                for col, header in enumerate(headers, 1):
                    cell = ws[f'{get_column_letter(col)}{current_row}']
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = border
                current_row += 1

                # Datos de productos
                for name, quantity, amount in top_products:
                    row_data = [
                        name or "Producto sin nombre",
                        quantity or 0,
                        format_currency(amount or 0)
                    ]

                    for col, value in enumerate(row_data, 1):
                        cell = ws[f'{get_column_letter(col)}{current_row}']
                        cell.value = value
                        cell.font = data_font
                        cell.border = border
                        if col in [2, 3]:  # Columnas numéricas
                            cell.alignment = Alignment(horizontal="right")
                        else:
                            cell.alignment = Alignment(horizontal="left")
                    current_row += 1

                current_row += 1

            # Distribución de pagos
            if payment_data:
                # Título de sección
                ws.merge_cells(f'A{current_row}:C{current_row}')
                payment_title = ws[f'A{current_row}']
                payment_title.value = "DISTRIBUCIÓN DE PAGOS"
                payment_title.font = subheader_font
                payment_title.fill = subheader_fill
                payment_title.alignment = Alignment(horizontal="center")
                current_row += 1

                # Encabezados
                headers = ['Método de Pago', 'Cantidad', 'Monto Total']
                for col, header in enumerate(headers, 1):
                    cell = ws[f'{get_column_letter(col)}{current_row}']
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = border
                current_row += 1

                # Datos de pagos
                for method, count, amount in payment_data:
                    if amount > 0:
                        row_data = [
                            method,
                            count,
                            format_currency(amount)
                        ]

                        for col, value in enumerate(row_data, 1):
                            cell = ws[f'{get_column_letter(col)}{current_row}']
                            cell.value = value
                            cell.font = data_font
                            cell.border = border
                            if col in [2, 3]:  # Columnas numéricas
                                cell.alignment = Alignment(horizontal="right")
                            else:
                                cell.alignment = Alignment(horizontal="left")
                        current_row += 1

            # Ajustar ancho de columnas
            for col in range(1, 7):
                ws.column_dimensions[get_column_letter(col)].width = 20

            # Guardar archivo
            wb.save(filename)

            QMessageBox.information(self.widget, "✅ Exportado", f"Reporte exportado exitosamente a:\n{filename}")

        except Exception as e:
            QMessageBox.critical(self.widget, "❌ Error", f"Error al exportar reporte:\n{str(e)}")

    def toggle_view_mode(self, is_detailed):
        """Mostrar/ocultar secciones según el modo de vista seleccionado"""
        # Las estadísticas principales siempre se muestran
        # Para vista detallada, mostrar todas las secciones
        # Para vista resumen, ocultar las secciones detalladas
        self.hourly_group.setVisible(is_detailed)
        self.payment_group.setVisible(is_detailed)

    def get_previous_period_dates(self, start_date, end_date):
        from datetime import datetime, timedelta
        try:
            current_start = datetime.strptime(start_date, "%Y-%m-%d")
            current_end = datetime.strptime(end_date, "%Y-%m-%d")
            period_days = (current_end - current_start).days + 1
            previous_end = current_start - timedelta(days=1)
            previous_start = previous_end - timedelta(days=period_days - 1)
            return previous_start.strftime("%Y-%m-%d"), previous_end.strftime("%Y-%m-%d")
        except Exception:
            return start_date, end_date
